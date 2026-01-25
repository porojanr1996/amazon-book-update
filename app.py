"""
Flask Web Application
Displays BSR charts and rankings
"""
from flask import Flask, render_template, jsonify, request
from flask_cors import CORS
from google_sheets_transposed import GoogleSheetsManager
from amazon_scraper import AmazonScraper
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
import config
import logging
import time
import re
from datetime import datetime, timedelta
import pytz
import os
from app.services.cache_service import invalidate_chart_cache as invalidate_chart_cache_service

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Install with: pip install openpyxl")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)


def parse_date(date_str):
    """
    Parse date string in various formats and return datetime object
    
    Supports:
    - YYYY-MM-DD (e.g., '2024-02-06')
    - M/D/YYYY (e.g., '2/6/2024', '02/06/2024')
    - MM/DD/YYYY (e.g., '02/06/2024')
    - M-D-YYYY (e.g., '2-6-2024')
    - Various other common formats
    
    Returns:
        datetime object or None if parsing fails
    """
    if not date_str:
        return None
    
    if isinstance(date_str, datetime):
        return date_str
    
    date_str = str(date_str).strip()
    
    if not date_str:
        return None
    
    # Try YYYY-MM-DD format first
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        pass
    
    # Try M/D/YYYY or MM/DD/YYYY format
    try:
        return datetime.strptime(date_str, '%m/%d/%Y')
    except ValueError:
        pass
    
    # Try M/D/YY format (2-digit year)
    try:
        parsed = datetime.strptime(date_str, '%m/%d/%y')
        # Convert 2-digit year to 4-digit (assume 2000s if < 50, else 1900s)
        if parsed.year < 50:
            parsed = parsed.replace(year=parsed.year + 2000)
        else:
            parsed = parsed.replace(year=parsed.year + 1900)
        return parsed
    except ValueError:
        pass
    
    # Try M-D-YYYY format
    try:
        return datetime.strptime(date_str, '%m-%d-%Y')
    except ValueError:
        pass
    
    # Try DD/MM/YYYY format (European)
    try:
        return datetime.strptime(date_str, '%d/%m/%Y')
    except ValueError:
        pass
    
    logger.warning(f"Could not parse date: {date_str}")
    return None


def normalize_date(date_str):
    """
    Normalize date string to YYYY-MM-DD format
    
    Returns:
        Normalized date string in YYYY-MM-DD format, or original string if parsing fails
    """
    dt = parse_date(date_str)
    if dt:
        return dt.strftime('%Y-%m-%d')
    return date_str

# Initialize Google Sheets manager
sheets_manager = None

# Cache for chart data (5 minute TTL)
_chart_data_cache = {}
_cache_timestamps = {}
CACHE_TTL = 300  # 5 minutes

# Cache for book covers (24 hour TTL)
_cover_cache = {}
_cover_cache_timestamps = {}
COVER_CACHE_TTL = 86400  # 24 hours

def get_sheets_manager():
    """Lazy initialization of sheets manager"""
    global sheets_manager
    if sheets_manager is None:
        sheets_manager = GoogleSheetsManager(
            config.GOOGLE_SHEETS_CREDENTIALS_PATH,
            config.GOOGLE_SHEETS_SPREADSHEET_ID
        )
    return sheets_manager

def get_cached_chart_data(time_range, worksheet=''):
    """Get cached chart data if available and not expired"""
    cache_key = f"{time_range}:{worksheet}"
    if cache_key in _chart_data_cache:
        cache_time = _cache_timestamps.get(cache_key, 0)
        if time.time() - cache_time < CACHE_TTL:
            logger.info(f"Returning cached chart data for {cache_key}")
            return _chart_data_cache[cache_key]
        else:
            # Cache expired, remove it
            del _chart_data_cache[cache_key]
            del _cache_timestamps[cache_key]
    return None

# Removed local invalidate_chart_cache() - now using Redis cache via app.services.cache_service

def set_cached_chart_data(time_range, worksheet='', data=None):
    """Cache chart data"""
    cache_key = f"{time_range}:{worksheet}"
    _chart_data_cache[cache_key] = data
    _cache_timestamps[cache_key] = time.time()


@app.route('/api/clear-cache')
def clear_cache():
    """Clear all caches (for debugging)"""
    global _chart_data_cache, _cache_timestamps, _cover_cache, _cover_cache_timestamps
    _chart_data_cache.clear()
    _cache_timestamps.clear()
    _cover_cache.clear()
    _cover_cache_timestamps.clear()
    logger.info("All caches cleared")
    return jsonify({'status': 'success', 'message': 'All caches cleared'})

@app.route('/')
def index():
    """Main dashboard page"""
    return render_template('index.html')


@app.route('/api/worksheets')
def get_worksheets():
    """Get all worksheet names from the Google Spreadsheet (excluding Sheet1 and Sheet3)."""
    try:
        manager = get_sheets_manager()
        all_worksheets = manager.get_all_worksheets()
        
        # Filter out Sheet1 and Sheet3
        excluded_sheets = ['Sheet1', 'Sheet3']
        worksheets = [ws for ws in all_worksheets if ws not in excluded_sheets]
        
        logger.info(f"Found {len(all_worksheets)} total worksheets: {all_worksheets}")
        logger.info(f"Returning {len(worksheets)} filtered worksheets: {worksheets}")
        return jsonify(worksheets)
    except Exception as e:
        logger.error(f"Error getting worksheets: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/books')
def get_books():
    """Get all books with BSR history for a specific worksheet"""
    try:
        worksheet_name = request.args.get('worksheet', '')
        manager = get_sheets_manager()
        
        # If no worksheet specified, get first worksheet as default
        if not worksheet_name:
            worksheets = manager.get_all_worksheets()
            if worksheets:
                worksheet_name = worksheets[0]
            else:
                worksheet_name = 'Crime Fiction - US'
        
        books_data = manager.get_bsr_history(worksheet_name=worksheet_name)
        return jsonify(books_data)
    except Exception as e:
        logger.error(f"Error getting books: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/rankings')
def get_rankings():
    """Get Top 50 rankings for a specific worksheet"""
    try:
        worksheet_name = request.args.get('worksheet', '')
        manager = get_sheets_manager()
        
        # If no worksheet specified, get first worksheet as default
        if not worksheet_name:
            worksheets = manager.get_all_worksheets()
            if worksheets:
                worksheet_name = worksheets[0]
            else:
                worksheet_name = 'Crime Fiction - US'
        
        logger.info(f"Getting rankings for worksheet: {worksheet_name}")
        books_data = manager.get_bsr_history(worksheet_name=worksheet_name)
        
        logger.info(f"Total books found in worksheet '{worksheet_name}': {len(books_data)}")
        
        # Separate books with and without BSR
        books_with_bsr = [b for b in books_data if b.get('current_bsr') is not None]
        books_without_bsr = [b for b in books_data if b.get('current_bsr') is None]
        
        logger.info(f"Books with BSR: {len(books_with_bsr)}, Books without BSR: {len(books_without_bsr)}")
        
        # Sort books with BSR by current BSR (lower is better)
        books_with_bsr.sort(key=lambda x: x['current_bsr'])
        
        # Combine: books with BSR first (sorted), then books without BSR
        # Return all books, not just top 50
        all_books = books_with_bsr + books_without_bsr
        
        logger.info(f"Returning all {len(all_books)} books from worksheet '{worksheet_name}'")
        
        # Extract cover images from cache (both local and Redis)
        # This ensures the endpoint responds quickly
        from app.services.cache_service import get_cached_cover
        
        for book in all_books:
            amazon_link = book.get('amazon_link')
            if amazon_link and not book.get('cover_image'):
                # First check local cache
                cache_key = amazon_link
                if cache_key in _cover_cache:
                    cache_time = _cover_cache_timestamps.get(cache_key, 0)
                    if time.time() - cache_time < COVER_CACHE_TTL:
                        book['cover_image'] = _cover_cache[cache_key]
                        continue
                
                # Then check Redis cache
                try:
                    redis_cover = get_cached_cover(amazon_link)
                    if redis_cover:
                        book['cover_image'] = redis_cover
                        # Also update local cache for faster access
                        _cover_cache[cache_key] = redis_cover
                        _cover_cache_timestamps[cache_key] = time.time()
                except Exception as e:
                    logger.debug(f"Could not get cover from Redis cache for {book.get('name', 'Unknown')}: {e}")
        
        # Log how many covers were found in cache
        covers_found = sum(1 for book in all_books if book.get('cover_image'))
        logger.info(f"Returning {len(all_books)} books, {covers_found} with cover images (from cache only)")
        
        return jsonify(all_books)
    except Exception as e:
        logger.error(f"Error getting rankings: {e}")
        return jsonify({'error': str(e)}), 500


def _enrich_books_with_excel_categories(books_data):
    """
    Enrich books with categories from Excel file
    Associates categories from categories.xlsx with books by Amazon link
    """
    if not OPENPYXL_AVAILABLE:
        return books_data
    
    excel_path = os.path.join(os.path.dirname(__file__), 'categories.xlsx')
    if not os.path.exists(excel_path):
        return books_data
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Create a mapping of Amazon link -> category
        link_to_category = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 4:
                amazon_link = str(row[2]).strip() if row[2] else ''
                category = str(row[3]).strip() if row[3] else ''
                
                if amazon_link and category:
                    # Filter out invalid categories
                    cat_clean = category.replace(',', '').replace(' ', '')
                    is_numeric = False
                    try:
                        float(cat_clean)
                        is_numeric = True
                    except ValueError:
                        pass
                    
                    if not is_numeric and category.lower() not in ['success', 'failed', 'error', 'none', '']:
                        # Normalize Amazon link for matching
                        clean_link = amazon_link.split('/ref')[0].split('?')[0].rstrip('/')
                        if not clean_link.endswith('/'):
                            clean_link += '/'
                        link_to_category[clean_link] = category
        
        # Enrich books with categories from Excel
        for book in books_data:
            amazon_link = book.get('amazon_link', '')
            if amazon_link:
                # Normalize link for matching
                clean_link = amazon_link.split('/ref')[0].split('?')[0].rstrip('/')
                if not clean_link.endswith('/'):
                    clean_link += '/'
                
                # If book doesn't have category or Excel has a better one, use Excel category
                if clean_link in link_to_category:
                    excel_category = link_to_category[clean_link]
                    current_category = book.get('category', '').strip()
                    
                    # Use Excel category if book doesn't have one, or if Excel category is more specific
                    if not current_category or (excel_category and excel_category.lower() not in ['kindle store', 'books']):
                        book['category'] = excel_category
                        logger.debug(f"Enriched {book.get('name', 'Unknown')} with category from Excel: {excel_category}")
        
        logger.info(f"Enriched {len([b for b in books_data if b.get('category')])} books with categories from Excel")
    except Exception as e:
        logger.warning(f"Could not enrich books with Excel categories: {e}")
    
    return books_data


@app.route('/api/chart-data')
def get_chart_data():
    """Get chart data for a specific time range and worksheet"""
    try:
        time_range = request.args.get('range', '30')  # days
        worksheet_name = request.args.get('worksheet', '')
        
        manager = get_sheets_manager()
        
        # If no worksheet specified, get first worksheet as default
        if not worksheet_name:
            worksheets = manager.get_all_worksheets()
            if worksheets:
                worksheet_name = worksheets[0]
            else:
                worksheet_name = 'Crime Fiction - US'
        
        start_time = time.time()
        
        logger.info(f"Chart data request: range={time_range}, worksheet={worksheet_name}")
        
        # Prevent browser caching for debugging
        response = None
        
        # Check cache first (include worksheet in cache key)
        cache_key = f"{time_range}:{worksheet_name}"
        cached_data = get_cached_chart_data(time_range, worksheet_name)
        if cached_data:
            logger.info(f"Returning cached data for range={time_range}, worksheet={worksheet_name}")
            return jsonify(cached_data)
        
        manager = get_sheets_manager()
        
        # Get average history directly from AVG column in Google Sheets
        avg_history = manager.get_avg_history(worksheet_name=worksheet_name)
        
        load_time = time.time() - start_time
        logger.info(f"Loaded {len(avg_history)} average values from AVG column in {load_time:.2f}s")
        logger.info(f"Total dates in Google Sheets: {len(avg_history)}")
        if avg_history:
            logger.info(f"Date range in Google Sheets: {avg_history[0].get('date', 'N/A')} to {avg_history[-1].get('date', 'N/A')}")
        
        # Process data for chart
        chart_data = {
            'dates': [],
            'average_bsr': [],
            'books': []
        }
        
        if not avg_history:
            logger.warning("No average history found")
            return jsonify(chart_data)
        
        # First, parse and normalize all dates
        parsed_history = []
        parse_failures = []
        for entry in avg_history:
            original_date = entry.get('date', '').strip()
            if not original_date:
                continue
            
            # Parse date
            entry_date = parse_date(original_date)
            if not entry_date:
                parse_failures.append(original_date)
                continue
            
            normalized_date = normalize_date(original_date)
            if normalized_date:
                parsed_entry = {
                    'date': normalized_date,
                    'original_date': original_date,
                    'average_bsr': entry.get('average_bsr'),
                    'parsed_date': entry_date
                }
                parsed_history.append(parsed_entry)
        
        if parse_failures:
            logger.warning(f"Failed to parse {len(parse_failures)} dates: {parse_failures[:10]}")
        
        # Log all dates from Google Sheets (including unparsed ones)
        logger.info(f"All dates from Google Sheets (first 10): {[e.get('date', 'N/A') for e in avg_history[:10]]}")
        logger.info(f"All dates from Google Sheets (last 10): {[e.get('date', 'N/A') for e in avg_history[-10:]]}")
        logger.info(f"Any 2025 dates in raw data: {[e.get('date', '') for e in avg_history if '2025' in str(e.get('date', ''))]}")
        
        # Sort by date to ensure correct order
        parsed_history.sort(key=lambda x: x['parsed_date'])
        
        # Find latest date after sorting
        latest_date = parsed_history[-1]['parsed_date'] if parsed_history else None
        
        # Log parsed dates
        logger.info(f"Parsed dates (first 10): {[e['date'] for e in parsed_history[:10]]}")
        logger.info(f"Parsed dates (last 10): {[e['date'] for e in parsed_history[-10:]]}")
        logger.info(f"Any 2025 dates in parsed data: {[e['date'] for e in parsed_history if '2025' in e['date']]}")
        
        if not parsed_history:
            logger.warning("No valid dates found in average history")
            return jsonify(chart_data)
        
        logger.info(f"Latest date in data: {latest_date.strftime('%Y-%m-%d')}")
        logger.info(f"Total parsed entries: {len(parsed_history)}")
        logger.info(f"Sample dates: {[e['date'] for e in parsed_history[:5]]}")
        logger.info(f"Last 5 dates: {[e['date'] for e in parsed_history[-5:]]}")
        
        # Calculate date range based on latest date, not current date
        if time_range == 'all':
            start_date = None
            logger.info("No date filtering for 'all' range")
        else:
            try:
                days = int(time_range)
                # Use latest_date as reference instead of datetime.now()
                start_date = latest_date - timedelta(days=days)
                logger.info(f"Filtering to last {days} days from {latest_date.strftime('%Y-%m-%d')} (start_date: {start_date.strftime('%Y-%m-%d')})")
            except ValueError:
                start_date = latest_date - timedelta(days=30)
                logger.info(f"Invalid range, defaulting to 30 days from {latest_date.strftime('%Y-%m-%d')} (start_date: {start_date.strftime('%Y-%m-%d')})")
        
        # Filter by date range
        filtered_history = []
        filtered_out_count = 0
        for entry in parsed_history:
            # Filter by date range
            if start_date is not None and entry['parsed_date'] < start_date:
                filtered_out_count += 1
                continue
            
            filtered_history.append({
                'date': entry['date'],
                'original_date': entry['original_date'],
                'average_bsr': entry['average_bsr']
            })
        
        logger.info(f"After filtering: {len(filtered_history)} entries kept, {filtered_out_count} filtered out")
        if filtered_history:
            logger.info(f"Filtered date range: {filtered_history[0]['date']} to {filtered_history[-1]['date']}")
        
        if not filtered_history:
            logger.warning("No average data found after filtering")
            return jsonify(chart_data)
        
        # Already sorted by parsed_date, but ensure normalized dates are also sorted
        filtered_history.sort(key=lambda x: parse_date(x['date']) or datetime.min)
        
        # Limit the number of dates for better chart performance (only for non-"all" ranges)
        # For "all" range, return all data points
        if time_range != 'all' and len(filtered_history) > 200:
            MAX_CHART_POINTS = 200
            logger.info(f"Limiting {len(filtered_history)} dates to {MAX_CHART_POINTS} for chart performance (range: {time_range})")
            step = max(1, len(filtered_history) // MAX_CHART_POINTS)
            filtered_history = filtered_history[::step][:MAX_CHART_POINTS]
            logger.info(f"After limiting: {len(filtered_history)} dates, first: {filtered_history[0]['date'] if filtered_history else 'N/A'}, last: {filtered_history[-1]['date'] if filtered_history else 'N/A'}")
        else:
            logger.info(f"Returning all {len(filtered_history)} dates for range '{time_range}'")
        
        # Extract dates and averages
        chart_data['dates'] = [entry['date'] for entry in filtered_history]
        chart_data['average_bsr'] = [entry['average_bsr'] for entry in filtered_history]
        
        logger.info(f"Chart will display {len(chart_data['dates'])} dates from AVG column")
        logger.info(f"Date range: {chart_data['dates'][0] if chart_data['dates'] else 'N/A'} to {chart_data['dates'][-1] if chart_data['dates'] else 'N/A'}")
        logger.info(f"Sample dates: {chart_data['dates'][:5] if len(chart_data['dates']) > 0 else 'N/A'}")
        logger.info(f"Sample averages: {chart_data['average_bsr'][:5] if len(chart_data['average_bsr']) > 0 else 'N/A'}")
        logger.info(f"All dates count: {len(chart_data['dates'])}, All averages count: {len(chart_data['average_bsr'])}")
        logger.info(f"Non-null averages: {sum(1 for x in chart_data['average_bsr'] if x is not None)}")
        
        # Get total books count for metadata
        manager = get_sheets_manager()
        books_data = manager.get_bsr_history(worksheet_name=worksheet_name)
        total_books_count = len(books_data)
        
        # Add metadata
        chart_data['total_books'] = total_books_count
        
        logger.info(f"Prepared {len(chart_data['average_bsr'])} average values from AVG column, {sum(1 for x in chart_data['average_bsr'] if x is not None)} non-null")
        
        # Add individual book data (optimized - only if needed, can be disabled for performance)
        # For now, skip individual book data to improve performance
        # Uncomment below if you need individual book lines on the chart
        # book_lookup = {}
        # for book in books_data:
        #     book_lookup[book['name']] = {}
        #     for entry in book.get('bsr_history', []):
        #         entry_normalized = normalize_date(entry['date'])
        #         if entry_normalized:
        #             book_lookup[book['name']][entry_normalized] = entry['bsr']
        # 
        # for book in books_data:
        #     book_series = {
        #         'name': book['name'],
        #         'author': book['author'],
        #         'data': []
        #     }
        #     book_data = book_lookup.get(book['name'], {})
        #     for display_date in chart_data['dates']:
        #         display_normalized = normalize_date(display_date)
        #         bsr_value = book_data.get(display_normalized)
        #         book_series['data'].append(bsr_value if bsr_value is not None else None)
        #     chart_data['books'].append(book_series)
        
        total_time = time.time() - start_time
        logger.info(f"Returning chart data: {len(chart_data['dates'])} dates, {len(chart_data['average_bsr'])} BSR values | Total time: {total_time:.2f}s")
        
        # Cache the result
        set_cached_chart_data(time_range, worksheet_name, chart_data)
        
        # Create response with no-cache headers to prevent browser caching
        response = jsonify(chart_data)
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        logger.error(f"Error getting chart data: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/categories')
def get_categories():
    """Get list of all categories"""
    try:
        manager = get_sheets_manager()
        
        # First try to get categories from Google Sheets
        books_data = manager.get_bsr_history()
        
        categories = set()
        all_category_values = []
        for book in books_data:
            cat = book.get('category', '').strip()
            all_category_values.append(cat)
            if cat:
                # Filter out numeric values (BSR values that might be in category column)
                # Check if it's a number (with or without commas)
                cat_clean = cat.replace(',', '').replace(' ', '')
                try:
                    # If it's a number, skip it (it's probably a BSR, not a category)
                    float(cat_clean)
                    logger.debug(f"Skipping numeric category value: {cat}")
                    continue
                except ValueError:
                    # Not a number, it's a valid category
                    categories.add(cat)
        
        # Also read categories from Excel file (combine with Google Sheets categories)
        if OPENPYXL_AVAILABLE:
            excel_path = os.path.join(os.path.dirname(__file__), 'categories.xlsx')
            if os.path.exists(excel_path):
                logger.info(f"No categories in Google Sheets, trying to read from {excel_path}")
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(excel_path)
                    ws = wb.active
                    
                    # Read categories from Excel (column D, starting from row 2)
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 4:
                            category = str(row[3]).strip() if row[3] else ''
                            if category and category.lower() not in ['success', 'failed', 'error', 'none', '']:
                                # Filter out numeric values
                                cat_clean = category.replace(',', '').replace(' ', '')
                                try:
                                    float(cat_clean)
                                    continue
                                except ValueError:
                                    categories.add(category)
                    
                    logger.info(f"Found {len(categories)} categories in Excel file")
                except Exception as e:
                    logger.warning(f"Could not read categories from Excel: {e}")
        
        # Log all category values for debugging
        logger.info(f"All category values from books: {all_category_values[:10]}... (showing first 10)")
        logger.info(f"Non-empty categories: {[c for c in all_category_values if c][:10]}")
        
        # Sort categories alphabetically
        sorted_categories = sorted(list(categories))
        logger.info(f"Found {len(sorted_categories)} valid categories: {sorted_categories}")
        
        # If no categories found, return empty array
        if not sorted_categories:
            logger.warning("No valid categories found. Categories may need to be extracted from Amazon.")
        
        return jsonify(sorted_categories)
    except Exception as e:
        logger.error(f"Error getting categories: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/extract-covers', methods=['POST'])
def extract_covers():
    """Extract cover images for all books in a worksheet (background task)"""
    try:
        data = request.get_json() or {}
        worksheet_name = data.get('worksheet', '')
        
        if not worksheet_name:
            worksheets = get_sheets_manager().get_all_worksheets()
            if worksheets:
                worksheet_name = worksheets[0]
            else:
                worksheet_name = 'Crime Fiction - US'
        
        logger.info(f"Starting cover extraction for worksheet: {worksheet_name}")
        
        # Run in background thread to avoid blocking
        def extract_covers_background():
            try:
                manager = get_sheets_manager()
                books = manager.get_all_books(worksheet_name=worksheet_name)
                
                from app.services.cache_service import get_cached_cover, set_cached_cover
                
                success_count = 0
                failure_count = 0
                
                for idx, book in enumerate(books):
                    amazon_link = book.get('amazon_link')
                    if not amazon_link:
                        continue
                    
                    # Skip if already cached
                    cached_cover = get_cached_cover(amazon_link)
                    if cached_cover:
                        logger.debug(f"Cover already cached for {book.get('name', 'Unknown')}")
                        continue
                    
                    # Add delay between requests to avoid rate limiting
                    # Use longer delay for cover extraction to be more conservative
                    if idx > 0:
                        delay = max(config.AMAZON_DELAY_BETWEEN_REQUESTS, 10.0)  # At least 10 seconds between covers
                        logger.debug(f"Waiting {delay}s before next cover extraction...")
                        time.sleep(delay)
                    
                    try:
                        logger.info(f"📸 Extracting cover for {book.get('name', 'Unknown')} ({idx + 1}/{len(books)})...")
                        scraper = AmazonScraper(
                            delay_between_requests=max(config.AMAZON_DELAY_BETWEEN_REQUESTS, 10.0),
                            retry_attempts=2  # 2 attempts for covers (more conservative)
                        )
                        
                        # Try with requests first
                        cover_url = scraper.extract_cover_image(amazon_link, use_playwright=False)
                        
                        # If failed and page might be blocked, wait longer before Playwright
                        if not cover_url:
                            logger.debug(f"First attempt failed, waiting before Playwright fallback...")
                            time.sleep(5)  # Extra delay before Playwright
                            cover_url = scraper.extract_cover_image(amazon_link, use_playwright=True)
                        
                        if cover_url:
                            # Clean up image URL for better quality
                            cover_url = re.sub(r'_SL\d+_', '_SL800_', cover_url)
                            cover_url = re.sub(r'\._AC_[^_]+_', '._AC_SL800_', cover_url)
                            if '_SL' not in cover_url:
                                cover_url = cover_url.replace('._AC_', '._AC_SL800_')
                            cover_url = re.sub(r'_SX\d+_', '_SX800_', cover_url)
                            
                            set_cached_cover(amazon_link, cover_url)
                            success_count += 1
                            logger.info(f"✓ Cover extracted for {book.get('name', 'Unknown')}: {cover_url[:80]}...")
                        else:
                            # Cache None but with shorter TTL to allow retry later
                            set_cached_cover(amazon_link, None)
                            failure_count += 1
                            logger.warning(f"✗ No cover found for {book.get('name', 'Unknown')} after all attempts")
                            # Extra delay after failure to avoid rate limiting
                            time.sleep(5)
                    except Exception as e:
                        logger.error(f"Error extracting cover for {book.get('name', 'Unknown')}: {e}")
                        failure_count += 1
                        try:
                            set_cached_cover(amazon_link, None)
                        except:
                            pass
                        # Extra delay after error to avoid rate limiting
                        time.sleep(5)
                
                logger.info(f"✅ Cover extraction completed: {success_count} success, {failure_count} failures")
            except Exception as e:
                logger.error(f"Error in background cover extraction: {e}", exc_info=True)
        
        # Start background thread
        import threading
        thread = threading.Thread(target=extract_covers_background, daemon=True)
        thread.start()
        
        return jsonify({
            'status': 'started',
            'message': f'Cover extraction started for worksheet "{worksheet_name}". Check logs for progress.',
            'worksheet': worksheet_name
        })
    except Exception as e:
        logger.error(f"Error starting cover extraction: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/update-bsr', methods=['POST'])
@app.route('/api/trigger-bsr-update', methods=['POST'])
def trigger_bsr_update():
    """
    Manually trigger BSR update for a specific worksheet (for testing)
    Note: This runs synchronously and may take time
    """
    try:
        from threading import Thread
        
        worksheet_name = request.json.get('worksheet', '') if request.is_json else request.form.get('worksheet', '')
        
        if not worksheet_name:
            # If no worksheet specified, update all worksheets
            logger.info("Manual BSR update triggered via API for all worksheets")
            # Run in background thread to avoid blocking
            thread = Thread(target=run_daily_bsr_update)
            thread.daemon = True
            thread.start()
            return jsonify({
                'status': 'started', 
                'message': 'BSR update started for all worksheets. Check logs for progress.'
            })
        else:
            logger.info(f"Manual BSR update triggered via API for worksheet: {worksheet_name}")
            # Run update for specific worksheet in background thread
            thread = Thread(target=run_worksheet_bsr_update, args=(worksheet_name,))
            thread.daemon = True
            thread.start()
            return jsonify({
                'status': 'started',
                'message': f'BSR update started for worksheet "{worksheet_name}". Check logs for progress.',
                'worksheet': worksheet_name
            })
    except Exception as e:
        logger.error(f"Error in manual BSR update: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


def run_worksheet_bsr_update(worksheet_name: str):
    """
    Update BSR for all books in a specific worksheet
    """
    logger.info("=" * 50)
    logger.info(f"Starting BSR update for worksheet: {worksheet_name}")
    logger.info(f"Time: {datetime.now(pytz.timezone('Europe/Bucharest'))}")
    logger.info("=" * 50)
    
    try:
        # Initialize components
        sheets_manager = get_sheets_manager()
        amazon_scraper = AmazonScraper(
            delay_between_requests=config.AMAZON_DELAY_BETWEEN_REQUESTS,
            retry_attempts=config.AMAZON_RETRY_ATTEMPTS
        )
        
        # Get all books from the specific worksheet
        books = sheets_manager.get_all_books(worksheet_name=worksheet_name)
        if not books:
            logger.warning(f"No books found in worksheet '{worksheet_name}'")
            return
        
        logger.info(f"Found {len(books)} books to process in worksheet '{worksheet_name}'")
        
        # Get today's row ONCE for this worksheet (avoid rate limiting)
        today_row = sheets_manager.get_today_row(worksheet_name=worksheet_name)
        logger.info(f"Today's BSR for {worksheet_name} will be written to row {today_row}")
        
        # Process books in parallel using threading
        success_count = 0
        failure_count = 0
        count_lock = Lock()
        
        def process_book(book):
            """Process a single book in a thread"""
            nonlocal success_count, failure_count
            
            # Create a new scraper instance for each thread (thread-safe)
            thread_scraper = AmazonScraper(
                delay_between_requests=config.AMAZON_DELAY_BETWEEN_REQUESTS,
                retry_attempts=config.AMAZON_RETRY_ATTEMPTS,
                use_tiered=True  # Enable tiered scraper
            )
            
            logger.info(f"Processing: {book['name']} by {book['author']} in {worksheet_name}")
            logger.info(f"Amazon URL: {book['amazon_link']}")
            
            try:
                # Extract BSR using tiered scraper (handles blocking automatically)
                bsr = thread_scraper.extract_bsr(book['amazon_link'])
                
                if bsr:
                    # Use the cached today_row (no additional API call)
                    sheets_manager.update_bsr(book['col'], today_row, bsr, worksheet_name=worksheet_name)
                    logger.info(f"✓ Successfully updated BSR: {bsr} for {book['name']} in {worksheet_name}")
                    with count_lock:
                        success_count += 1
                    return True
                else:
                    logger.warning(f"✗ Could not extract BSR for {book['name']} in {worksheet_name}")
                    with count_lock:
                        failure_count += 1
                    return False
            except Exception as e:
                logger.error(f"✗ Error processing {book['name']} in {worksheet_name}: {e}", exc_info=True)
                with count_lock:
                    failure_count += 1
                return False
        
        # Use ThreadPoolExecutor for parallel processing
        max_workers = min(config.AMAZON_MAX_WORKERS, len(books))
        logger.info(f"Processing {len(books)} books with {max_workers} threads in parallel for {worksheet_name}...")
        
        process_start_time = time.time()
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            for idx, book in enumerate(books, 1):
                if idx > 1:
                    time.sleep(config.AMAZON_DELAY_BETWEEN_REQUESTS / max_workers)
                future = executor.submit(process_book, book)
                futures[future] = book
            
            for future in as_completed(futures):
                book = futures[future]
                try:
                    future.result()
                except Exception as e:
                    logger.error(f"Fatal error processing {book['name']} in {worksheet_name}: {e}", exc_info=True)
        
        elapsed_time = time.time() - process_start_time
        logger.info(f"Processing for {worksheet_name} completed in {elapsed_time:.2f} seconds ({len(books)/elapsed_time:.2f} books/sec)")
        
        # Calculate and update average BSR for today's row for this worksheet
        try:
            logger.info(f"Calculating average BSR for today in worksheet: {worksheet_name}...")
            current_today_row = sheets_manager.get_today_row(worksheet_name=worksheet_name)
            sheets_manager.calculate_and_update_average(current_today_row, worksheet_name=worksheet_name)
            logger.info(f"✓ Average BSR calculated and updated successfully for {worksheet_name}")
            
            # Flush all buffered updates to Google Sheets
            logger.info(f"Flushing buffered updates to Google Sheets for {worksheet_name}...")
            sheets_manager.flush_batch_updates(worksheet_name=worksheet_name)
            logger.info(f"✓ All updates flushed to Google Sheets for {worksheet_name}")
            
            # Invalidate chart cache for this specific worksheet
            logger.info(f"Invalidating chart cache for worksheet: {worksheet_name}")
            try:
                invalidate_chart_cache_service(worksheet_name=worksheet_name)
                logger.info(f"✓ Chart cache invalidated for {worksheet_name}")
            except Exception as e:
                logger.error(f"✗ Error invalidating chart cache: {e}", exc_info=True)
        except Exception as e:
            logger.error(f"✗ Error calculating average BSR or flushing updates for {worksheet_name}: {e}", exc_info=True)
        
        logger.info("=" * 50)
        logger.info(f"BSR update completed for worksheet: {worksheet_name}")
        logger.info(f"Success: {success_count}")
        logger.info(f"Failures: {failure_count}")
        logger.info("=" * 50)
        
    except Exception as e:
        logger.error(f"Error in worksheet BSR update: {e}", exc_info=True)


@app.route('/api/extract-categories', methods=['POST'])
def extract_categories():
    """
    Extract categories from Amazon for all books and save to Excel file
    This only extracts categories, doesn't update BSR
    Saves results to categories.xlsx in the project root
    """
    try:
        if not OPENPYXL_AVAILABLE:
            return jsonify({
                'status': 'error',
                'message': 'openpyxl not installed. Install with: pip install openpyxl'
            }), 500
        
        logger.info("=" * 50)
        logger.info("Starting category extraction from Amazon")
        logger.info("=" * 50)
        
        sheets_manager = get_sheets_manager()
        # Use longer delay for category extraction to avoid rate limiting
        amazon_scraper = AmazonScraper(
            delay_between_requests=6.0,  # 6 seconds between requests for category extraction
            retry_attempts=config.AMAZON_RETRY_ATTEMPTS
        )
        
        books = sheets_manager.get_all_books()
        if not books:
            return jsonify({'status': 'error', 'message': 'No books found'}), 400
        
        logger.info(f"Found {len(books)} books to process")
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Categories"
        
        # Add headers
        ws.append(['Book Name', 'Author', 'Amazon Link', 'Category', 'Status'])
        
        success_count = 0
        failure_count = 0
        
        for idx, book in enumerate(books, 1):
            logger.info(f"Processing {idx}/{len(books)}: {book['name']}")
            
            # Clean URL - remove ref parameters and trailing slashes
            clean_url = book['amazon_link'].split('/ref')[0].split('?')[0].rstrip('/')
            if not clean_url.endswith('/'):
                clean_url += '/'
            
            logger.info(f"Amazon URL: {clean_url}")
            try:
                # Use Selenium for category extraction (more reliable for JavaScript-rendered content)
                book_info = amazon_scraper.extract_book_info(clean_url, use_selenium=True)
                if book_info:
                    logger.debug(f"Book info extracted: {list(book_info.keys())}")
                    if 'category' in book_info:
                        logger.debug(f"Category found in book_info: {book_info['category']}")
                    else:
                        logger.debug("No category in book_info")
                else:
                    logger.warning(f"No book info extracted for {book['name']}")
                    book_info = {}
                
                if book_info.get('category'):
                    category = book_info['category']
                    ws.append([
                        book['name'],
                        book.get('author', ''),
                        book['amazon_link'],
                        category,
                        'Success'
                    ])
                    logger.info(f"✓ Extracted category for {book['name']}: {category}")
                    success_count += 1
                else:
                    reason = 'Failed - Category not found'
                    if not book_info:
                        reason = 'Failed - Could not extract book info'
                    ws.append([
                        book['name'],
                        book.get('author', ''),
                        book['amazon_link'],
                        '',
                        reason
                    ])
                    logger.warning(f"✗ Could not extract category for {book['name']}: {reason}")
                    failure_count += 1
                
                # Additional delay after each request to avoid rate limiting
                if idx < len(books):  # Don't delay after last book
                    logger.debug(f"Waiting 6 seconds before next request...")
                    time.sleep(6.0)
                    
            except Exception as e:
                ws.append([
                    book['name'],
                    book.get('author', ''),
                    book['amazon_link'],
                    '',
                    f'Error: {str(e)}'
                ])
                logger.error(f"✗ Error processing {book['name']}: {e}")
                failure_count += 1
                # Delay even on error to avoid rate limiting
                if idx < len(books):
                    logger.debug(f"Waiting 6 seconds after error before next request...")
                    time.sleep(6.0)
        
        # Save Excel file
        excel_filename = 'categories.xlsx'
        excel_path = os.path.join(os.path.dirname(__file__), excel_filename)
        wb.save(excel_path)
        
        logger.info("=" * 50)
        logger.info(f"Category extraction completed: {success_count} success, {failure_count} failures")
        logger.info(f"Results saved to: {excel_path}")
        logger.info("=" * 50)
        
        return jsonify({
            'status': 'success',
            'message': f'Category extraction completed: {success_count} success, {failure_count} failures',
            'success_count': success_count,
            'failure_count': failure_count,
            'excel_file': excel_filename,
            'excel_path': excel_path
        })
    except Exception as e:
        logger.error(f"Error extracting categories: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/update-categories-from-excel', methods=['POST'])
def update_categories_from_excel():
    """
    Update categories in Google Sheets from categories.xlsx file
    Reads categories from Excel and updates them in Google Sheets Row 4
    """
    try:
        if not OPENPYXL_AVAILABLE:
            return jsonify({
                'status': 'error',
                'message': 'openpyxl not installed. Install with: pip install openpyxl'
            }), 500
        
        excel_path = os.path.join(os.path.dirname(__file__), 'categories.xlsx')
        if not os.path.exists(excel_path):
            return jsonify({
                'status': 'error',
                'message': f'categories.xlsx not found at {excel_path}'
            }), 404
        
        logger.info("=" * 50)
        logger.info("Starting category update from Excel to Google Sheets")
        logger.info("=" * 50)
        
        # Read categories from Excel
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Get all books from Google Sheets to match by name/author/link
        sheets_manager = get_sheets_manager()
        books = sheets_manager.get_all_books()
        
        # Create lookup by Amazon link (most reliable)
        books_by_link = {book['amazon_link']: book for book in books}
        
        updated_count = 0
        not_found_count = 0
        
        # Read categories from Excel (column D, starting from row 2)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) >= 4:
                book_name = str(row[0]).strip() if row[0] else ''
                author = str(row[1]).strip() if row[1] else ''
                amazon_link = str(row[2]).strip() if row[2] else ''
                category = str(row[3]).strip() if row[3] else ''
                status = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                
                # Skip if no category or if status indicates failure
                if not category or category.lower() in ['success', 'failed', 'error', 'none', '']:
                    continue
                
                # Skip numeric values
                cat_clean = category.replace(',', '').replace(' ', '')
                try:
                    float(cat_clean)
                    continue
                except ValueError:
                    pass
                
                # Find matching book by Amazon link
                if amazon_link in books_by_link:
                    book = books_by_link[amazon_link]
                    try:
                        sheets_manager.update_category(book['col'], category)
                        logger.info(f"✓ Updated category for {book['name']}: {category}")
                        updated_count += 1
                    except Exception as e:
                        logger.error(f"✗ Error updating category for {book['name']}: {e}")
                else:
                    logger.warning(f"✗ Book not found in Google Sheets: {book_name} ({amazon_link})")
                    not_found_count += 1
        
        logger.info("=" * 50)
        logger.info(f"Category update completed: {updated_count} updated, {not_found_count} not found")
        logger.info("=" * 50)
        
        return jsonify({
            'status': 'success',
            'message': f'Categories updated: {updated_count} updated, {not_found_count} not found',
            'updated_count': updated_count,
            'not_found_count': not_found_count
        })
    except Exception as e:
        logger.error(f"Error updating categories from Excel: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/scheduler-status')
def get_scheduler_status():
    """Get scheduler status and next run time"""
    try:
        jobs = scheduler.get_jobs()
        job_info = []
        for job in jobs:
            next_run = None
            if scheduler.running:
                try:
                    next_run = job.next_run_time
                    if next_run:
                        next_run = next_run.isoformat()
                except AttributeError:
                    pass
            
            job_info.append({
                'id': job.id,
                'name': job.name,
                'next_run_time': next_run,
                'trigger': str(job.trigger)
            })
        
        return jsonify({
            'scheduler_running': scheduler.running,
            'jobs': job_info
        })
    except Exception as e:
        logger.error(f"Error getting scheduler status: {e}")
        return jsonify({'error': str(e)}), 500


def run_daily_bsr_update():
    """
    Function to update BSR ratings daily
    This function is called by the scheduler at 10:01 AM Bucharest time
    """
    logger.info("=" * 50)
    logger.info("Starting scheduled daily BSR update")
    logger.info(f"Time: {datetime.now(pytz.timezone('Europe/Bucharest'))}")
    logger.info("=" * 50)
    
    try:
        # Initialize components
        sheets_manager = get_sheets_manager()
        amazon_scraper = AmazonScraper(
            delay_between_requests=config.AMAZON_DELAY_BETWEEN_REQUESTS,
            retry_attempts=config.AMAZON_RETRY_ATTEMPTS
        )
        
        # Get all books from all worksheets
        books = []
        worksheets = sheets_manager.get_all_worksheets()
        logger.info(f"Found {len(worksheets)} worksheets: {worksheets}")
        for worksheet_name in worksheets:
            try:
                worksheet_books = sheets_manager.get_all_books(worksheet_name=worksheet_name)
                # Add worksheet name to each book
                for book in worksheet_books:
                    book['worksheet'] = worksheet_name
                books.extend(worksheet_books)
                logger.info(f"Loaded {len(worksheet_books)} books from worksheet '{worksheet_name}'")
            except Exception as e:
                logger.error(f"Error loading books from worksheet '{worksheet_name}': {e}")
                continue
        
        if not books:
            logger.warning("No books found in Google Sheets")
            return
        
        logger.info(f"Found {len(books)} books to process from {len(worksheets)} worksheets")
        
        # Process books in parallel using threading
        success_count = 0
        failure_count = 0
        count_lock = Lock()
        
        def process_book(book):
            """Process a single book in a thread"""
            nonlocal success_count, failure_count  # Declare nonlocal at the start of the function
            
            # Create a new scraper instance for each thread (thread-safe)
            thread_scraper = AmazonScraper(
                delay_between_requests=config.AMAZON_DELAY_BETWEEN_REQUESTS,
                retry_attempts=config.AMAZON_RETRY_ATTEMPTS
            )
            
            logger.info(f"Processing: {book['name']} by {book['author']}")
            logger.info(f"Amazon URL: {book['amazon_link']}")
            
            try:
                # Extract BSR
                bsr = thread_scraper.extract_bsr(book['amazon_link'])
                
                # Extract category if not already set
                current_category = book.get('category', '').strip()
                if not current_category:
                    # Try to extract category from Amazon
                    logger.info(f"Category not set, extracting from Amazon...")
                    book_info = thread_scraper.extract_book_info(book['amazon_link'])
                    if book_info and book_info.get('category'):
                        category = book_info['category']
                        sheets_manager.update_category(book['col'], category)
                        logger.info(f"✓ Updated category: {category}")
                    else:
                        logger.warning(f"Could not extract category from Amazon")
                
                if bsr:
                    # Get today's row again (in case it changed)
                    worksheet_name = book.get('worksheet', 'Crime Fiction - US')
                    today_row = sheets_manager.get_today_row(worksheet_name=worksheet_name)
                    sheets_manager.update_bsr(book['col'], today_row, bsr, worksheet_name=worksheet_name)
                    logger.info(f"✓ Successfully updated BSR for {book['name']} in worksheet '{worksheet_name}': {bsr}")
                    with count_lock:
                        success_count += 1
                    return True
                else:
                    logger.warning(f"✗ Could not extract BSR for {book['name']}")
                    with count_lock:
                        failure_count += 1
                    return False
            except Exception as e:
                logger.error(f"✗ Error processing {book['name']}: {e}")
                with count_lock:
                    failure_count += 1
                return False
        
        # Use ThreadPoolExecutor for parallel processing
        max_workers = min(config.AMAZON_MAX_WORKERS, len(books))  # Configurable or number of books
        logger.info(f"Processing {len(books)} books with {max_workers} threads in parallel...")
        
        start_time = time.time()
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit tasks with delay between starts to avoid rate limiting
            futures = {}
            for idx, book in enumerate(books, 1):
                # Staggered start - delay between thread starts
                if idx > 1:
                    time.sleep(config.AMAZON_DELAY_BETWEEN_REQUESTS / max_workers)
                future = executor.submit(process_book, book)
                futures[future] = book
            
            # Process results as they complete
            for future in as_completed(futures):
                book = futures[future]
                try:
                    future.result()
                except Exception as e:
                    logger.error(f"Fatal error processing {book['name']}: {e}")
        
        elapsed_time = time.time() - start_time
        logger.info(f"Processing completed in {elapsed_time:.2f} seconds ({len(books)/elapsed_time:.2f} books/sec)")
        
        # Calculate and update average BSR for today's row (for each worksheet)
        # Get unique worksheets from books
        worksheets = set(book.get('worksheet', 'Crime Fiction - US') for book in books)
        for worksheet_name in worksheets:
            try:
                logger.info(f"Calculating average BSR for today in worksheet: {worksheet_name}...")
                today_row = sheets_manager.get_today_row(worksheet_name=worksheet_name)
                sheets_manager.calculate_and_update_average(today_row, worksheet_name=worksheet_name)
                logger.info(f"✓ Average BSR calculated and updated successfully for {worksheet_name}")
                
                # Flush all buffered updates to Google Sheets for this worksheet
                logger.info(f"Flushing buffered updates to Google Sheets for {worksheet_name}...")
                sheets_manager.flush_batch_updates(worksheet_name=worksheet_name)
                logger.info(f"✓ All updates flushed to Google Sheets for {worksheet_name}")
            except Exception as e:
                logger.error(f"✗ Error calculating average BSR or flushing updates for {worksheet_name}: {e}")
        
        # Invalidate chart cache so new average appears immediately
        invalidate_chart_cache_service()
        
        # Summary
        logger.info("=" * 50)
        logger.info("Daily BSR update completed")
        logger.info(f"Success: {success_count}")
        logger.info(f"Failures: {failure_count}")
        logger.info("=" * 50)
        
    except Exception as e:
        logger.error(f"Fatal error in daily BSR update: {e}", exc_info=True)


# Initialize scheduler
scheduler = BackgroundScheduler(timezone=pytz.timezone('Europe/Bucharest'))
scheduler.add_job(
    func=run_daily_bsr_update,
    trigger=CronTrigger(hour=10, minute=0, timezone=pytz.timezone('Europe/Bucharest')),
    id='daily_bsr_update',
    name='Daily BSR Update at 10:00 AM Bucharest time',
    replace_existing=True
)


if __name__ == '__main__':
    # Start scheduler
    try:
        scheduler.start()
        logger.info("Scheduler started successfully")
        logger.info("Daily BSR update scheduled for 10:01 AM Bucharest time")
    except Exception as e:
        logger.error(f"Failed to start scheduler: {e}", exc_info=True)
    
    # Run Flask app
    try:
        app.run(
            host=config.FLASK_HOST,
            port=config.FLASK_PORT,
            debug=(config.FLASK_ENV == 'development')
        )
    except (KeyboardInterrupt, SystemExit):
        logger.info("Shutting down scheduler...")
        scheduler.shutdown()

